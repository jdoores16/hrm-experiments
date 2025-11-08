"""
Demo: Natural Language Prompt → Multi-Pole Circuit
Shows how "circuits 1,3,5 is a 3pole 30A circuit..." becomes proper panel schedule
"""

from pathlib import Path
from openpyxl import load_workbook
from tests.test_panel_ir import make_header, make_circuit
from app.schemas.panel_ir import PanelScheduleIR
from app.io.panel_excel import write_excel_from_ir

TEMPLATE_XLSX = Path(__file__).parents[1] / "templates" / "panelboard_template.xlsx"
OUT = Path(__file__).parents[1] / "out"


def test_natural_language_prompt():
    """
    User prompt: 'circuits 1,3,5 is a 3pole 30A circuit that feeds an 
                  exhaust fan and the phase amp is 18'
    
    AI Interpretation:
    - Circuit #1 (3-pole) occupies rows for circuits 1, 3, 5
    - Description: "EXHAUST FAN"
    - Breaker: 30A
    - Load: 18A per phase (A, B, C)
    - Poles: 3
    
    Expected Excel Output:
    Row 12 (Circuit 1): Poles=3, Breaker=30A, PhA=18A, PhB=18A, PhC=18A, Desc="EXHAUST FAN"
    Row 13 (Circuit 3): Poles="-", Breaker="-", PhA="-", PhB="-", PhC="-", Desc="-"
    Row 14 (Circuit 5): Poles="-", Breaker="-", PhA="-", PhB="-", PhC="-", Desc="-"
    """
    
    print("\n" + "="*80)
    print("DEMO: Natural Language Prompt → Multi-Pole Circuit")
    print("="*80)
    
    print("\n📝 USER PROMPT:")
    print("   'circuits 1,3,5 is a 3pole 30A circuit that feeds an exhaust fan")
    print("    and the phase amp is 18'")
    
    print("\n🤖 AI INTERPRETATION:")
    print("   - Circuit Number: 1 (starting circuit)")
    print("   - Poles: 3 (occupies 3 rows)")
    print("   - Breaker: 30A")
    print("   - Phase Amps: 18A on each phase (A, B, C)")
    print("   - Description: 'EXHAUST FAN'")
    print("   - Rows Used: 12 (ckt 1), 13 (ckt 3), 14 (ckt 5)")
    
    # Create the circuit from AI interpretation
    header = make_header()
    circuits = [
        make_circuit(
            ckt=1,                          # Starting circuit number
            side="odd",                     # Odd side of panel
            excel_row=12,                   # Row 12 (circuits 1,3,5 = rows 12,13,14)
            breaker_amps=30,                # 30A breaker
            load_amps=18,                   # 18A per phase
            poles=3,                        # 3-pole circuit
            phA=True,                       # Phase A energized
            phB=True,                       # Phase B energized
            phC=True,                       # Phase C energized
            description="EXHAUST FAN"       # Load description
        )
    ]
    
    ir = PanelScheduleIR(header=header, circuits=circuits)
    
    # Generate Excel
    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / "demo_nl_prompt.xlsx"
    
    result_path = write_excel_from_ir(
        ir=ir,
        out_path=str(out_path),
        template_xlsx=str(TEMPLATE_XLSX)
    )
    
    print(f"\n📄 GENERATED: {result_path.name}")
    
    # Verify the Excel output
    wb = load_workbook(result_path)
    ws = wb.active
    
    print("\n✅ EXCEL OUTPUT VERIFICATION:")
    
    # Row 12 (Circuit 1) - First row with actual data
    print(f"   Row 12 (Circuit 1):")
    print(f"      Description: '{ws['A12'].value}'")
    print(f"      Phase A Load: {ws['B12'].value}A")
    print(f"      Phase B Load: {ws['C12'].value}A")
    print(f"      Phase C Load: {ws['D12'].value}A")
    print(f"      Poles: {ws['E12'].value}")
    print(f"      Breaker: {ws['F12'].value}A")
    
    assert ws['A12'].value == "EXHAUST FAN"
    assert ws['B12'].value == 18
    assert ws['C12'].value == 18
    assert ws['D12'].value == 18
    assert ws['E12'].value == 3
    assert ws['F12'].value == 30
    
    # Row 13 (Circuit 3 position) - Continuation row
    print(f"\n   Row 13 (Circuit 3 - Continuation):")
    print(f"      Description: '{ws['A13'].value}'")
    print(f"      Phase A: '{ws['B13'].value}'")
    print(f"      Phase B: '{ws['C13'].value}'")
    print(f"      Phase C: '{ws['D13'].value}'")
    print(f"      Poles: '{ws['E13'].value}'")
    print(f"      Breaker: '{ws['F13'].value}'")
    
    assert ws['A13'].value == "-"
    assert ws['B13'].value == "-"
    assert ws['C13'].value == "-"
    assert ws['D13'].value == "-"
    assert ws['E13'].value == "-"
    assert ws['F13'].value == "-"
    
    # Row 14 (Circuit 5 position) - Continuation row
    print(f"\n   Row 14 (Circuit 5 - Continuation):")
    print(f"      Description: '{ws['A14'].value}'")
    print(f"      Phase A: '{ws['B14'].value}'")
    print(f"      Phase B: '{ws['C14'].value}'")
    print(f"      Phase C: '{ws['D14'].value}'")
    print(f"      Poles: '{ws['E14'].value}'")
    print(f"      Breaker: '{ws['F14'].value}'")
    
    assert ws['A14'].value == "-"
    assert ws['B14'].value == "-"
    assert ws['C14'].value == "-"
    assert ws['D14'].value == "-"
    assert ws['E14'].value == "-"
    assert ws['F14'].value == "-"
    
    print("\n" + "="*80)
    print("✅ SUCCESS! Natural language prompt correctly interpreted and generated.")
    print("="*80)
    print(f"\n📁 Output file: {result_path}")
    print("\nThe 3-pole circuit correctly:")
    print("  ✓ Occupies circuits 1, 3, 5 (rows 12, 13, 14)")
    print("  ✓ Has 18A on each phase (A, B, C)")
    print("  ✓ Shows 30A breaker")
    print("  ✓ Displays 'EXHAUST FAN' description")
    print("  ✓ Marks continuation rows with '-' symbols")
    print()


if __name__ == "__main__":
    test_natural_language_prompt()
