"""
Test multi-pole circuit handling in panel schedules.
Verifies that 2-pole and 3-pole circuits create continuation rows with "-" symbols.
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from tests.test_panel_ir import make_header, make_circuit
from app.schemas.panel_ir import PanelScheduleIR
from app.io.panel_excel import write_excel_from_ir

TEMPLATE_XLSX = Path(__file__).parents[1] / "templates" / "panelboard_template.xlsx"
OUT = Path(__file__).parents[1] / "out"


def test_single_pole_circuit():
    """Test that 1-pole circuits occupy a single row"""
    header = make_header()
    circuits = [
        make_circuit(ckt=1, side="odd", excel_row=12, breaker_amps=20, load_amps=15, poles=1, phA=True)
    ]
    
    ir = PanelScheduleIR(header=header, circuits=circuits)
    
    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / "test_1pole.xlsx"
    
    result_path = write_excel_from_ir(
        ir=ir,
        out_path=str(out_path),
        template_xlsx=str(TEMPLATE_XLSX)
    )
    
    # Verify the file was created
    wb = load_workbook(result_path)
    ws = wb.active
    
    # Row 12 should have the circuit data (circuit 1)
    assert ws['F12'].value == 20  # Breaker amps
    assert ws['E12'].value == 1  # Poles
    assert ws['B12'].value == 15  # Phase A load
    assert ws['A12'].value == "CIRCUIT 1"  # Description
    
    # Row 13 should NOT have continuation "-" (next circuit or blank)
    assert ws['E13'].value != "-"
    
    print("✓ Single-pole circuit test passed")


def test_two_pole_circuit():
    """Test that 2-pole circuits occupy two rows with continuation"""
    header = make_header()
    circuits = [
        make_circuit(ckt=1, side="odd", excel_row=12, breaker_amps=30, load_amps=25, poles=2, 
                    phA=True, phB=True, description="2-POLE LOAD")
    ]
    
    ir = PanelScheduleIR(header=header, circuits=circuits)
    
    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / "test_2pole.xlsx"
    
    result_path = write_excel_from_ir(
        ir=ir,
        out_path=str(out_path),
        template_xlsx=str(TEMPLATE_XLSX)
    )
    
    wb = load_workbook(result_path)
    ws = wb.active
    
    # Row 12 (first row) - should have actual data
    assert ws['F12'].value == 30  # Breaker amps
    assert ws['E12'].value == 2  # Poles
    assert ws['B12'].value == 25  # Phase A load
    assert ws['C12'].value == 25  # Phase B load
    assert ws['A12'].value == "2-POLE LOAD"  # Description
    
    # Row 13 (continuation row) - should have "-" symbols
    assert ws['A13'].value == "-"  # Description continuation
    assert ws['F13'].value == "-"  # Breaker continuation
    assert ws['E13'].value == "-"  # Poles continuation
    assert ws['B13'].value == "-"  # Phase A continuation
    assert ws['C13'].value == "-"  # Phase B continuation
    assert ws['D13'].value == "-"  # Phase C continuation
    
    print("✓ Two-pole circuit test passed")


def test_three_pole_circuit():
    """Test that 3-pole circuits occupy three rows with continuations"""
    header = make_header()
    circuits = [
        make_circuit(ckt=1, side="odd", excel_row=12, breaker_amps=60, load_amps=50, poles=3,
                    phA=True, phB=True, phC=True, description="3-POLE MOTOR")
    ]
    
    ir = PanelScheduleIR(header=header, circuits=circuits)
    
    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / "test_3pole.xlsx"
    
    result_path = write_excel_from_ir(
        ir=ir,
        out_path=str(out_path),
        template_xlsx=str(TEMPLATE_XLSX)
    )
    
    wb = load_workbook(result_path)
    ws = wb.active
    
    # Row 12 (first row) - should have actual data
    assert ws['F12'].value == 60  # Breaker amps
    assert ws['E12'].value == 3  # Poles
    assert ws['B12'].value == 50  # Phase A load
    assert ws['C12'].value == 50  # Phase B load
    assert ws['D12'].value == 50  # Phase C load
    assert ws['A12'].value == "3-POLE MOTOR"  # Description
    
    # Row 13 (first continuation) - should have "-" symbols
    assert ws['A13'].value == "-"
    assert ws['F13'].value == "-"
    assert ws['E13'].value == "-"
    assert ws['B13'].value == "-"
    assert ws['C13'].value == "-"
    assert ws['D13'].value == "-"
    
    # Row 14 (second continuation) - should have "-" symbols
    assert ws['A14'].value == "-"
    assert ws['F14'].value == "-"
    assert ws['E14'].value == "-"
    assert ws['B14'].value == "-"
    assert ws['C14'].value == "-"
    assert ws['D14'].value == "-"
    
    print("✓ Three-pole circuit test passed")


def test_mixed_pole_circuits():
    """Test panel with mix of 1-pole, 2-pole, and 3-pole circuits"""
    header = make_header()
    circuits = [
        make_circuit(ckt=1, poles=1, phA=True),
        make_circuit(ckt=7, poles=2, phA=True, phB=True),
        make_circuit(ckt=13, poles=3, phA=True, phB=True, phC=True),
    ]
    
    ir = PanelScheduleIR(header=header, circuits=circuits)
    
    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / "test_mixed.xlsx"
    
    result_path = write_excel_from_ir(
        ir=ir,
        out_path=str(out_path),
        template_xlsx=str(TEMPLATE_XLSX)
    )
    
    wb = load_workbook(result_path)
    ws = wb.active
    
    # Circuit 1 (1-pole, row 12 = 11 + (1+1)//2)
    assert ws['E12'].value == 1
    
    # Circuit 7 (2-pole, row 15-16 = 11 + (7+1)//2)
    assert ws['E15'].value == 2
    assert ws['E16'].value == "-"
    
    # Circuit 13 (3-pole, row 18-20 = 11 + (13+1)//2)
    assert ws['E18'].value == 3
    assert ws['E19'].value == "-"
    assert ws['E20'].value == "-"
    
    print("✓ Mixed pole circuit test passed")
